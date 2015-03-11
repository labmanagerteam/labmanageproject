# -*- coding: utf-8 -*-
__author__ = 'wlw'

from django.shortcuts import render
from django.http import HttpResponseRedirect

from labmanageproject.user_manage import *
from labmanageproject.my_decorator import *
from labmanageproject.error_code import *
from labmanageproject.utility import *
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook


def login(request):
    if 'my_user' in request.session:
        return render(request, "home.html", locals())

    login_error = False
    if request.method == 'POST':
        form = login_form(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            uname = check_password(cd['uid'], cd['password'])
            # print uname
            if uname:
                [perm_list, identity_list] = get_perm_list(cd['uid'])
                my_user = {
                    'perm_list': perm_list,
                    'uname': uname,
                    'uid': cd['uid'],
                    'identity': identity_list,
                    'password': cd['password'],
                }
                request.session['my_user'] = my_user
                return render(request, 'home.html', locals())
            else:
                login_error = True
    else:
        form = login_form()
    # return render_to_response('login.html', locals(), context_instance=RequestContext(request, c))
    return render(request, 'login.html', locals())


def logout(request):
    del request.session['my_user']
    return HttpResponseRedirect('/')


def add_user_view(request):
    add_list = [
        {
            'val': '0',
            'name': 'student'
        },
        {
            'val': '1',
            'name': 'teacher'
        }
    ]
    return render(request, "add_user2.html", locals())


def wrap_row(row):
    def to_str(a):
        if type(a) != type(u""):
            return str(a).split('.')[0]
        else:
            return a

    l = []
    print "wrap_row"
    print row
    try:
        for cell in row:
            print type(cell.internal_value)
            l.append(to_str(cell.internal_value))
            print cell.internal_value
        print "wrap_row ended"
    except Exception, e:
        print e.message
    return l


student_distribute_list = ['uid', 'uname', 'password', 'card_number', 'grade', 'did']


@check_post_form(student_distribute_list)
def add_one_student_view(request):
    value_list = get_post(request, student_distribute_list)
    error_code = add_one_student_action(*value_list)
    if error_code:
        error_message = generate_error_message(error_code)
        return create_json_return({'result': 'error', 'msg': error_message})
    else:
        return create_json_return({'result': 'success'})


def add_student_list_view(request):
    f = request.FILES['user_list']

    def add_student_list_handler(wb):
        ws = wb.get_sheet_by_name(name='Sheet1')
        first = True
        print ws
        student_list = []
        for row in ws.iter_rows():
            if first:
                first = False
                continue
            else:
                student_list.append(wrap_row(row))
                # print row
                wrap_row(row)
                print "one_student added"
        add_student_list_action(student_list)

    print "add_student_list_view"
    try:
        do_with_xlsx(add_student_list_handler, f.read())
        return create_json_return({'result': 'success'})
    except MyListException, e:
        error_message = generate_error_message(e.error_code)
        return create_json_return({'result': 'error', 'msg': error_message})
    except Exception, e:
        return create_json_return({'result': 'error', 'msg': e.message})


teacher_distribute_list = ['uid', 'uname', 'password', 'lcid', 'card_number']


@check_post_form(teacher_distribute_list)
def add_one_teacher_view(request, is_admin=False):
    value_list = get_post(request, teacher_distribute_list)
    error_code = add_one_teacher_action(*value_list, is_admin=is_admin)
    if error_code:
        error_message = generate_error_message(error_code)
        return create_json_return({'result': 'error', 'msg': error_message})
    else:
        return create_json_return({'result': 'success'})


def add_teacher_list_view(request):
    f = request.FILES['user_list']

    def test_handler(wb):
        ws = wb.get_sheet_by_name(name='Sheet1')
        for row in ws.iter_rows():
            for cell in row:
                print cell.internal_value

    def add_teacher_list_handler(wb):

        ws = wb.get_sheet_by_name(name='Sheet1')
        first = True
        teacher_list = []
        for row in ws.iter_rows():
            if first:
                first = False
                continue
            else:
                teacher_list.append(wrap_row(row))
        add_teacher_list_action(teacher_list)

    print "add_teacher_list_view"
    try:
        do_with_xlsx(add_teacher_list_handler, f.read())
    except MyBaseException, e:
        error_message = generate_error_message(e.error_code)
        return create_json_return({'result': 'error', 'msg': error_message})
    return create_json_return({'result': 'success'})


def add_view(request):
    add_list = [
        {
            'val': '2',
            'name': '实验室'
        },
        {
            'val': '3',
            'name': '实验中心'
        },
        {
            'val': '4',
            'name': '院系'
        },
        {
            'val': '5',
            'name': '实验中心管理员'
        }
    ]
    return render(request, "add_user2.html", locals())


lab_distribution_list = ['lid', 'lname', 'lcid', 'lnumber']


@check_post_form(lab_distribution_list)
def add_one_lab_view(request):
    l = get_post(request, lab_distribution_list)
    add_one_lab_action_action(*l)
    return create_json_return({'result': 'success'})


lab_center_dist_list = ['lcid', 'lcname']


@check_post_form(lab_center_dist_list)
def add_one_lab_center_view(request):
    print('add_one_lab_center')
    l = get_post(request, lab_center_dist_list)
    add_one_lab_center_action(*l)
    return create_json_return({'result': 'success'})


department_dist_list = ['did', 'dname']


@check_post_form(department_dist_list)
def add_one_department_view(request):
    add_one_department_action(*get_post(request, department_dist_list))
    return create_json_return({'result': 'success'})


def add_one_admin_view(request):
    return add_one_teacher_view(request, True)


def get_all_lab_center_admin_view(request):
    result = get_all_lab_center_admin_action()
    print result
    return render(request, "get_all_lab_center_admin.html", locals())


@check_post_form({'delid'})
def delete_one_lab_center_admin_view(request):
    [uid] = get_post(request, ['delid'])
    delete_one_admin_action(uid)
    return create_json_return({'result': 'success'})


def delete_one_item_view_factory(delete_action):
    @check_post_form({'delid'})
    def delete(request):
        [delete_id] = get_post(request, ['delid'])
        delete_action(delete_id)
        return create_json_return({'result': 'success'})

    return delete


delete_one_lab_center_view = delete_one_item_view_factory(delete_one_lab_center_action)
delete_one_lab_view = delete_one_item_view_factory(delete_one_lab_action)


def get_all_lab_center_view(request):
    lab_center_list = get_all_lab_center_action()
    return render(request, "get_all_lab_center.html", locals())


def get_one_lab_center_detail_view(request, lcid):
    one_lab_center = {
        'lcid': lcid,
        'lcname': get_lab_center_table(**{'lcid': lcid})[0][1]
    }

    lab_list = get_lab_by_lcid(lcid)

    return render(request, "get_lab_center_detail.html", locals())


def do_with_xlsx(handler, f):
    temp = NamedTemporaryFile(suffix='.xlsx')
    try:
        temp.write(f)
        print u"temp_name:%s" % temp.name
        name = u"%s" % temp.name
        temp.seek(0)
        wb = load_workbook(filename=name, use_iterators=True)
        handler(wb)
    finally:
        temp.close()


def add_list_view_factory(name, do_function):
    def add_list(request):
        f = request.FILES['user_list']

        def add_list_handler(wb):
            ws = wb.get_sheet_by_name(name='Sheet1')
            first = True
            print ws
            student_list = []
            for row in ws.iter_rows():
                if first:
                    first = False
                    continue
                else:
                    student_list.append(wrap_row(row))
                    # print row
                    wrap_row(row)
                    print "one_student added"
            do_function(student_list)

        print name
        try:
            do_with_xlsx(add_list_handler, f.read())
        except MyListException, e:
            error_message = generate_error_message(e.error_code, e.num)
            return create_json_return({'result': 'error', 'msg': error_message})
        return create_json_return({'result': 'success'})

    return add_list


add_lab_center_list_view = add_list_view_factory('add_lab_center_list', add_lab_center_list_action)
add_lab_list_view = add_list_view_factory('add_lab_list', add_lab_list_action)
add_department_list_view = add_list_view_factory('add_department_list', add_department_list_action)
add_admin_list_view = add_list_view_factory('add_admin_list_view', add_admin_list_action)


@check_post_form(['uid', 'new_password'])
def change_admin_password_view(request):
    [uid, new_password] = get_post(request, ['uid', 'new_password'])
    try:
        change_password_action(uid, new_password)
        return create_json_return({'result': 'success'})
    except MyBaseException, e:
        return create_json_return({'result': 'error', 'msg': generate_error_message(e.error_code)})